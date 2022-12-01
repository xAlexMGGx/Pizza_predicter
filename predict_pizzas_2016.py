import pandas as pd
import re
from word2number import w2n
import re
from xml.dom import minidom
import xml.etree.ElementTree as ET
import generate_xml_2016 as gx
import matplotlib.pyplot as plt
import generate_pdf_2016 as gp
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side


# DATA OBTENTION
def extract():
    pizzas = pd.read_csv('ORIGINALS/pizzas.csv')
    pizza_types = pd.read_csv('ORIGINALS/pizza_types.csv', encoding='latin-1')
    orders = pd.read_csv('ORIGINALS/orders_2016.csv', sep=';')
    order_details = pd.read_csv('ORIGINALS/order_details_2016.csv', sep=';', encoding='latin-1')
    data_dictionary = pd.read_csv('ORIGINALS/data_dictionary.csv')
    return pizzas, pizza_types, orders, order_details, data_dictionary


# DATA ANALYSIS
def anomalies(order_details, pizzas):
    anomaly = False
    for pizza in order_details['pizza_id']:
        if pizza not in pizzas['pizza_id'].unique():
            print(f'Pizza ID "{pizza}" not recognized')
            anomaly = True
    if not anomaly:
        print('No anomalies found')


def analyze():
    data = extract()
    df_name = ['pizzas', 'pizza_types', 'orders', 'order_details', 'data_dictionary']
    for idx, df in enumerate(data):
        print('\nDataframe:', df_name[idx])
        print(df.head())
        print('\nColumn dtypes:')
        print(df.dtypes)
        print('\nNumber of null values:')
        print(df.isnull().sum())
    print('\nAnomalies:')
    anomalies(data[3], data[0])


# DATA TRANFORMATION
# First, we will get rid of the null values in the orders and order_details dataframe
def clean_data(ordered_pizzas):
    ordered_pizzas.pop('time')
    ordered_pizzas.pop('order_details_id')
    ordered_pizzas = ordered_pizzas.dropna()
    ordered_pizzas = ordered_pizzas[ordered_pizzas['pizza_id'] != 'nan']
    ordered_pizzas = ordered_pizzas.reset_index(drop=True)
    return ordered_pizzas


# Then we will transform the data to the correct format
def transform_data(orders, order_details):
    ordered_pizzas = pd.merge(orders, order_details, on='order_id')
    ordered_pizzas['date'] = [control_date(x) for x in ordered_pizzas['date']]
    # We will use regex to write pizza_id with the correct format (e.g. pizza_flv_size)
    ordered_pizzas['pizza_id'] = [re.sub('[ -]', '_', str(x)) for x in ordered_pizzas['pizza_id']]
    # There are some values that are misswritten in the pizza_id column, so we will correct them
    ordered_pizzas['pizza_id'] = [re.sub('@', 'a', str(x)) for x in ordered_pizzas['pizza_id']]
    ordered_pizzas['pizza_id'] = [re.sub('3', 'e', str(x)) for x in ordered_pizzas['pizza_id']]
    ordered_pizzas['pizza_id'] = [re.sub('0', 'o', str(x)) for x in ordered_pizzas['pizza_id']]
    # We will use the word2number library to transform the string number to integer
    ordered_pizzas['quantity'] = [control_w2n(x) for x in ordered_pizzas['quantity']]
    ordered_pizzas = clean_data(ordered_pizzas)
    return ordered_pizzas.astype({'quantity': 'int64'})


def control_w2n(x):
    try:
        return w2n.word_to_num(str(x))
    except:
        return


def control_date(x):
    try:
        if re.findall('\A\d\d-\d\d-\d\d', x):
            return pd.to_datetime(x, format='%d-%m-%y %H:%M:%S')
        else:
            return pd.to_datetime(float(x)+3600, unit='s')
    except:
        return pd.to_datetime(x)


# We create a csv with the pizzas ordered in each order, instead of having each pizza in a different row
def csv_orders(ordered_pizzas):
    pizza_order = {}
    for i in ordered_pizzas['order_id']:
        if i not in pizza_order:
            pizza_order[i] = []
    for i in range(len(ordered_pizzas)):
        for _ in range(ordered_pizzas['quantity'][i]):
            pizza_order[ordered_pizzas['order_id'][i]].append(ordered_pizzas['pizza_id'][i])
    ordered_pizzas['pizzas'] = [pizza_order[x] for x in ordered_pizzas['order_id']]
    ordered_pizzas.pop('quantity')
    ordered_pizzas.pop('pizza_id')
    ordered_pizzas = (ordered_pizzas.drop_duplicates(subset=['order_id'])).reset_index(drop=True)
    ordered_pizzas.to_csv('TRANSFORMED/ordered_pizzas_2016.csv', index=False)
    return ordered_pizzas


# This function adds to the ordered_pizzas dataframe the day of the week of each order
def csv_with_days(ordered_pizzas):
    dates = []
    days = {0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday', 5: 'Saturday', 6: 'Sunday'}
    for date in ordered_pizzas['date']:
        dates.append(days[date.weekday()])
    ordered_pizzas['day'] = dates
    ordered_pizzas.to_csv('TRANSFORMED/ordered_pizzas_2016.csv', index=False)
    return ordered_pizzas


# This function creates a csv with all the ingredients and the amount of each one for each day of the week (and total week)
def csv_ingredients(pizza_types):
    all_ingredients = []
    for ingredients in pizza_types['ingredients']:
        for ingredient in ingredients.split(', '):
            ingredient = re.sub('ï¿½','', ingredient)
            if ingredient not in all_ingredients:
                all_ingredients.append(ingredient)
    ingredients_df = pd.DataFrame({'ingredient': all_ingredients, 'Monday': [0 for _ in range(len(all_ingredients))], 'Tuesday': [0 for i in range(len(all_ingredients))], 'Wednesday': [0 for i in range(len(all_ingredients))], 'Thursday': [0 for i in range(len(all_ingredients))], 'Friday': [0 for i in range(len(all_ingredients))], 'Saturday': [0 for i in range(len(all_ingredients))], 'Sunday': [0 for i in range(len(all_ingredients))], 'Total': [0 for i in range(len(all_ingredients))]})
    ingredients_df.to_csv('TRANSFORMED/ingredients_2016.csv', index=False)
    return ingredients_df


# We create a function that returns the pizza flavour and size of a pizza
def search_pizza(pizza_id, pizzas):
    pizza = pizzas[pizzas['pizza_id'] == pizza_id]
    return pizza['pizza_type_id'].values[0], pizza['size'].values[0]


# This function adds to the pizzas csv the amount of pizzas ordered each day of the week
# We calculate this by pizza, where we use the size as a factor 
def create_csv_with_pizzas_per_day(ordered_pizzas, pizza_types, pizzas_data, date):
    lst = [0 for _ in range(len(pizza_types))]
    weigths = {'S': 1, 'M': 1.5, 'L': 2, 'XL': 2.5, 'XXL': 3}
    pizza_counts = {'Monday': lst.copy(), 'Tuesday': lst.copy(), 'Wednesday': lst.copy(), 'Thursday': lst.copy(), 'Friday': lst.copy(), 'Saturday': lst.copy(), 'Sunday': lst.copy()}
    for _, order in ordered_pizzas.iterrows():
        if pd.to_datetime(order['date']) >= date:
            break
        else:
            day = order['day']
            pizzas = order['pizzas']
            for pizza in pizzas:
                pizza_flavour, size = search_pizza(pizza, pizzas_data)
                ind = pizza_types[pizza_types['pizza_type_id'] == pizza_flavour].index.values[0]
                pizza_counts[day][ind] += 1*weigths[size]
    for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
        pizza_types[day] = pizza_counts[day]
    pizza_types.to_csv('TRANSFORMED/pizza_counts_2016.csv', index=False)
    return pizza_types


# This function does the same as the previous one for each week (in order to add it to the report)
def count_pizzas_per_week(i, pizza_counts, ordered_pizzas, pizza_types, pizzas_data, date0, date1):
    lst = [0 for _ in range(len(pizza_types))]
    weigths = {'S': 1, 'M': 1.5, 'L': 2, 'XL': 2.5, 'XXL': 3}
    for _, order in ordered_pizzas.iterrows():
        if pd.to_datetime(order['date']) < date0:
            continue
        elif pd.to_datetime(order['date']) >= date1:
            break
        else:
            pizzas = order['pizzas']
            for pizza in pizzas:
                pizza_flavour, size = search_pizza(pizza, pizzas_data)
                ind = pizza_types[pizza_types['pizza_type_id'] == pizza_flavour].index.values[0]
                lst[ind] += 1*weigths[size]
    pizza_counts[f'Week {i}'] = lst
    return pizza_counts


def create_df_with_pizzas_per_week(pizza_types, ordered_pizzas, pizzas_data):
    date0 = pd.to_datetime('2016-01-01')
    date1 = pd.to_datetime('2016-01-08')
    i = 0
    pizza_counts = pd.DataFrame()
    pizza_counts.index = pizza_types['pizza_type_id']
    pizza_counts.index.name = None
    while date0 < pd.to_datetime('2016-12-31'):
        pizza_counts = count_pizzas_per_week(i, pizza_counts, ordered_pizzas, pizza_types, pizzas_data, date0, date1)
        date0 = date1
        date1 += pd.Timedelta(days=7)
        i += 1
    for week in pizza_counts.columns:
        pizza_counts[week] = pizza_counts[week].astype(int)
    pizza_counts.to_csv('TRANSFORMED/pizza_counts_per_week_2016.csv')
    return pizza_counts


# This function calculates the amount of ingredients needed a specific day of the week
def ingredients_quantity(day, pizza_types, pizza_type_id, days_difference):
    aux = (pizza_types[pizza_types['pizza_type_id'] == pizza_type_id][day].values[0])
    return aux*7/days_difference


# This function predicts the ingredients needed for the following week
def predict(pizza_types, ingredients_df, days_difference):
    for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
        for ingredients in pizza_types['ingredients']:
            pizza_type_id = pizza_types[pizza_types['ingredients'] == ingredients]['pizza_type_id'].values[0]
            for ingredient in ingredients.split(', '): 
                ingredient = re.sub('ï¿½','', ingredient)
                ind = ingredients_df[ingredients_df['ingredient'] == ingredient].index.values[0]   
                prediction = ingredients_quantity(day, pizza_types, pizza_type_id, days_difference)
                ingredients_df.loc[ind,[day]] += prediction
                ingredients_df.loc[ind,['Total']] += prediction
    ingredients_df.to_csv('TRANSFORMED/ingredients_2016.csv', index=False)
    return ingredients_df


# XML
def prettify(elem):
    """Return a pretty-printed XML string for the Element.
    """
    rough_string = ET.tostring(elem, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="    ")


def create_xml(ingredients_df):
    data = {}
    data['prediction_per_ingredient'] = {}
    for i in range(len(ingredients_df)):
        data['prediction_per_ingredient'][re.sub(' ', '_', (ingredients_df['ingredient'][i]))] = ingredients_df['Total'][i]
    root = ET.Element('prediction')
    sub = ET.SubElement(root, 'prediction_per_ingredient')
    for key, value in data['prediction_per_ingredient'].items():
        ET.SubElement(sub, key).text = str(value)

    with open('data_report_2016.xml', 'a') as f:
        f.write(prettify(root))


# IMAGES GENERATION FOR PDF
def create_images(ingredients_df):
    # Stacked bar chart
    ings_df = ingredients_df.copy()
    plt.figure(figsize=(40,20))
    colors = ['red', 'green', 'blue', 'yellow', 'orange', 'purple', 'pink']
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    bot = None
    for i, day in enumerate(days):
        plt.bar(ings_df['ingredient'], ings_df[day], bottom=bot, color=colors[i])
        if bot is None:
            bot = ings_df[day]
        else:
            bot += ings_df[day]
    plt.xticks(rotation=90, fontsize=20)
    plt.yticks(fontsize=20)
    plt.title('Ingredients needed for the week', fontsize=30)
    plt.xlabel('Ingredients', fontsize=20)
    plt.ylabel('Quantity', fontsize=20)
    plt.legend(days, fontsize=20)
    plt.savefig('IMAGES/ingredients_2016.png', bbox_inches='tight', transparent=False)


# EXCEL
def create_excel(ingredients_df, pizza_counts):
    # Generate excel file
    with pd.ExcelWriter('data_report_2016.xlsx', engine='openpyxl') as writer:
        ingredients_df.to_excel(writer, startrow=6, startcol=2, sheet_name='Ingredients report', index=False)
        pizza_counts.to_excel(writer, startrow=6, startcol=2, sheet_name='Pizzas report', index=True)
    wb = load_workbook('data_report_2016.xlsx')
    ws1 = wb['Ingredients report']
    ws2 = wb['Pizzas report']

    # Sheet 1 - Ingredients report
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['F'].width = 12
    ws1['C3'] = 'Ingredients needed for the week'
    ws1['C3'].border = Border(bottom=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    ws1.merge_cells('C3:K3')
    ws1['C3'].font = Font(size=20, bold=True)
    ws1['C3'].alignment = Alignment(horizontal='center')
    col_0, col_1 = wb.active.min_column, wb.active.max_column - 1
    row_0, row_1 = wb.active.min_row, wb.active.max_row
    # Create a stacked bar chart
    bcht = BarChart()
    data = Reference(ws1, min_col=col_0+1, min_row=row_0, max_col=col_1, max_row=row_1)
    cats = Reference(ws1, min_col=col_0, min_row=row_0+1, max_col=col_0, max_row=row_1)
    bcht.grouping = 'stacked'
    bcht.overlap = 100
    bcht.add_data(data, titles_from_data=True)
    bcht.set_categories(cats)
    bcht.title = 'Ingredients needed for the week'
    bcht.style = 2
    bcht.y_axis.title = 'Quantity'
    bcht.x_axis.title = 'Ingredients'
    bcht.width = 40
    bcht.height = 20
    ws1.add_chart(bcht, "N3")

    # Sheet 2 - Pizzas report
    ws2.column_dimensions['C'].width = 20
    # Align all column C to the left
    for i in range(1, ws2.max_row+1):
        ws2.cell(row=i, column=3).alignment = Alignment(horizontal='left')
    ws2['C3'] = 'Pizza types report'
    ws2['C3'].border = Border(bottom=Side(border_style='thin', color='000000'),
                              top=Side(border_style='thin', color='000000'),
                              left=Side(border_style='thin', color='000000'),
                              right=Side(border_style='thin', color='000000'))
    ws2.merge_cells('C3:K3')
    ws2['C3'].font = Font(size=20, bold=True)
    ws2['C3'].alignment = Alignment(horizontal='center')
    ws2['C7'] = 'Pizza type'
    ws2['C7'].font = Font(bold=True)
    # Put border to cell C7
    ws2['C7'].border = Border(left=Side(border_style='thin', color='000000'),
                              right=Side(border_style='thin', color='000000'),
                              top=Side(border_style='thin', color='000000'),
                              bottom=Side(border_style='thin', color='000000'))

    wb.save('data_report_2016.xlsx')


# MAIN
def main():
    # PREDICTION
    date = pd.to_datetime('2016-12-31', format='%Y-%m-%d')
    days_difference = (date - pd.to_datetime('2016-01-01', format='%Y-%m-%d')).days
    pizzas, pizza_types, orders, order_details, data_dictionary = extract()
    ordered_pizzas = transform_data(orders, order_details)
    ordered_pizzas = csv_orders(ordered_pizzas)
    ordered_pizzas = csv_with_days(ordered_pizzas)
    ordered_pizzas = ordered_pizzas.sort_values(by=['date'])
    pizza_types = create_csv_with_pizzas_per_day(ordered_pizzas, pizza_types, pizzas, date)
    pizza_counts = create_df_with_pizzas_per_week(pizza_types, ordered_pizzas, pizzas)
    ingredients_df = csv_ingredients(pizza_types)
    ingredients_df = predict(pizza_types, ingredients_df, days_difference)
    ingredients_df = ingredients_df.astype({'Monday': 'int', 'Tuesday': 'int', 'Wednesday': 'int', 'Thursday': 'int', 'Friday': 'int', 'Saturday': 'int', 'Sunday': 'int', 'Total': 'int'})
    print('This are the ingredients that you need to buy for the week:')
    print(ingredients_df)

    # XML
    gx.main()
    create_xml(ingredients_df)

    # PDF
    create_images(ingredients_df)
    gp.main()

    # EXCEL
    create_excel(ingredients_df, pizza_counts)


if __name__ == '__main__':
    main()

